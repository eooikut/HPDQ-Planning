from flask import Blueprint, render_template, request, redirect, url_for
from uuid import uuid4
import os, shutil
import pandas as pd
from sqlalchemy import text
# üü¢ Import c√°c h√†m util (kh√¥ng ƒë·ªïi)
from storage_utils import load_metadata, get_lsx_by_id, save_metadata, update_metadata
# üü¢ Import h√†m x·ª≠ l√Ω file input
from XuLyDuLieuNhapVao import process_create_lsx
from datetime import datetime
import re
from auth.decorator import permission_required
# üü¢ Engine SQL Server d√πng chung
from db import engine        # <- b·∫°n ƒë√£ t·∫°o engine trong db_utils
import pandas as pd
import openpyxl
import io
import sqlalchemy
from openpyxl.utils import get_column_letter 
from flask import (
    request, 
    send_file, 
    jsonify, 
    session, 
    Blueprint # Ho·∫∑c 'Flask' n·∫øu b·∫°n d√πng 1 file
)
lsx_bp = Blueprint("lsx", __name__)


@lsx_bp.route("/danhsach")
@permission_required('manage_lsx')
def danhsach_lsx():
    """
    Li·ªát k√™ to√†n b·ªô LSX t·ª´ metadata (file JSON).
    """
    metadata = load_metadata()
    lsx_list = [d for d in metadata if d.get("type") == "lsx"]  # ‚úÖ ch·ªâ l·∫•y LSX
    if not lsx_list:
        return render_template("xem_theo_ngay.html")
    return render_template("danhsach.html", datasets=lsx_list)


@lsx_bp.route("/xem/<lsx_id>")
@permission_required('manage_lsx')
def xem_thong_tin_lsx(lsx_id):
    """
    Xem chi ti·∫øt 1 LSX.
    """
    selected = get_lsx_by_id(lsx_id)
    if not selected:
        return "Kh√¥ng t√¨m th·∫•y LSX", 404
    return render_template("chitietlsx.html", lsx=selected)


@lsx_bp.route("/edit/<lsx_id>", methods=["GET", "POST"])
@permission_required('manage_lsx')
def edit_lsx(lsx_id):
    """
    Ch·ªânh s·ª≠a t√™n LSX.
    """
    entry = get_lsx_by_id(lsx_id)
    if not entry:
        return "Kh√¥ng t√¨m th·∫•y LSX", 404

    if request.method == "POST":
        new_name = request.form.get("name", entry["name"])
        entry["name"] = new_name
        update_metadata(lsx_id, entry)     # c·∫≠p nh·∫≠t JSON
        return redirect(url_for("lsx.danhsach_lsx"))

    return render_template("edit_lsx.html", entry=entry)


@lsx_bp.route("/delete/<lsx_id>", methods=["POST"])
@permission_required('manage_lsx')
def delete_lsx(lsx_id):
    metadata = load_metadata()
    entry = get_lsx_by_id(lsx_id)
    if not entry:
        return "Kh√¥ng t√¨m th·∫•y LSX", 404

    folder = os.path.dirname(entry["lsx"])
    try:
        if os.path.exists(folder):
            shutil.rmtree(folder)
    except Exception as e:
        return f"L·ªói khi x√≥a th∆∞ m·ª•c: {e}", 500

    # ‚úÖ Gi·ªØ l·∫°i c√°c b·∫£n ghi kh√°c lo·∫°i ho·∫∑c LSX kh√°c ID
    metadata = [d for d in metadata if not (d.get("type") == "lsx" and d.get("id") == lsx_id)]
    save_metadata(metadata)

    # --- X√≥a trong DB SQL Server ---
    tables = ["ton_kho", "report", "lsx"]
    with engine.begin() as conn:
        for table in tables:
            try:
                stmt = text(f"DELETE FROM {table} WHERE lsx_id = :lsx_id")
                result = conn.execute(stmt, {"lsx_id": lsx_id})
                print(f"Deleted {result.rowcount} rows from {table}")
            except Exception as e:
                print(f"L·ªói khi x√≥a b·∫£ng {table}: {e}")
                pass

    # --- ƒêi·ªÅu h∆∞·ªõng ---
    remaining_lsx = [d for d in metadata if d.get("type") == "lsx"]
    if not remaining_lsx:
        return render_template("no_data.html")
    return redirect(url_for("lsx.danhsach_lsx"))
@lsx_bp.route('/api/get-lsx-data', methods=['GET'])
@permission_required('manage_lsx') # Th√™m decorator ƒë·ªÉ b·∫£o v·ªá API
def get_lsx_data():
    try:
        import numpy as np

        sql = "SELECT * FROM LenhSanXuat_ChiTiet ORDER BY STT"
        
        # D√πng `engine` b·∫°n ƒë√£ t·∫°o
        with engine.connect() as conn:
            df = pd.read_sql(sql, conn)

        # === S·ª¨A L·ªñI: Thay th·∫ø NaN b·∫±ng None ===
        # Chuy·ªÉn ƒë·ªïi t·∫•t c·∫£ c√°c gi√° tr·ªã NaN trong DataFrame th√†nh None c·ªßa Python
        # tr∆∞·ªõc khi g·ª≠i v·ªÅ client. Khi jsonify(None) s·∫Ω t·∫°o ra chu·ªói 'null'
        # h·ª£p l·ªá trong JSON, m√† JavaScript c√≥ th·ªÉ hi·ªÉu ƒë∆∞·ª£c.
        df = df.replace({np.nan: None})

        data_json = df.to_dict(orient='records')
        return jsonify(data_json)
        
    except Exception as e:
        print(f"L·ªói get_lsx_data: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500

# =================================================================
# API 2: L∆∞u d·ªØ li·ªáu (POST)
# =================================================================
@lsx_bp.route('/api/save-lsx-data', methods=['POST'])
@permission_required('manage_lsx') # Th√™m d√≤ng n√†y ƒë·ªÉ ki·ªÉm tra quy·ªÅn
def save_lsx_data():
    data_rows = request.get_json()
    if not data_rows:
        return jsonify({"status": "error", "message": "Kh√¥ng c√≥ d·ªØ li·ªáu"}), 400

    try:
        with engine.begin() as conn: # B·∫Øt ƒë·∫ßu m·ªôt transaction
            # 1. L·∫•y danh s√°ch ID t·ª´ frontend v√† ID t·ª´ database
            frontend_ids = {row.get('ID') for row in data_rows if row.get('ID') is not None}
            
            # L·∫•y t·∫•t c·∫£ ID hi·ªán c√≥ trong b·∫£ng
            result = conn.execute(text("SELECT ID FROM LenhSanXuat_ChiTiet"))
            db_ids = {row[0] for row in result}

            # 2. X√°c ƒë·ªãnh c√°c ID c·∫ßn x√≥a
            ids_to_delete = db_ids - frontend_ids
            if ids_to_delete:
                # Chuy·ªÉn set th√†nh list ƒë·ªÉ d√πng trong c√¢u l·ªánh IN
                delete_list = list(ids_to_delete)
                # T·∫°o placeholders cho c√¢u l·ªánh IN
                placeholders = ', '.join([f':id_{i}' for i in range(len(delete_list))])
                delete_stmt = text(f"DELETE FROM LenhSanXuat_ChiTiet WHERE ID IN ({placeholders})")
                
                # T·∫°o dictionary cho c√°c tham s·ªë
                params = {f'id_{i}': val for i, val in enumerate(delete_list)}
                conn.execute(delete_stmt, params)
                print(f"--- DEBUG: ƒê√£ x√≥a {len(ids_to_delete)} d√≤ng v·ªõi ID: {ids_to_delete} ---")

            # 3. L·∫∑p qua d·ªØ li·ªáu t·ª´ frontend ƒë·ªÉ INSERT ho·∫∑c UPDATE
            for row in data_rows:
                row_id = row.get('ID')
                params = {
                        "STT": row.get('STT'), "ThoiGianSX": row.get('ThoiGianSX'), "KichCo": row.get('KichCo'),
                        "MacThep": row.get('MacThep'), "SanLuong_1A": row.get('SanLuong_1A'), "SanLuong_1B": row.get('SanLuong_1B'),
                        "SanLuong_YeuCau_Cuon": row.get('SanLuong_YeuCau_Cuon'), "DungSai": row.get('DungSai'),
                        "CoTinh_GHC": row.get('CoTinh_GHC'), "CoTinh_GHB": row.get('CoTinh_GHB'), "CoTinh_GianDai": row.get('CoTinh_GianDai'),
                        "CoTinh_DoCung": row.get('CoTinh_DoCung'), "Phoi_MacPhoi": row.get('Phoi_MacPhoi'),
                        "Phoi_KichThuoc": row.get('Phoi_KichThuoc'), "YeuCauDacBiet": row.get('YeuCauDacBiet'),
                        "OrderNumber": row.get('OrderNumber'), "Batch": row.get('Batch'), "KL_Cuon": row.get('KL_Cuon'),
                        "MucDichSuDung": row.get('MucDichSuDung'), "KhachHang": row.get('KhachHang'),
                        "DotSX": row.get('DotSX'), # Th√™m DotSX
                        "NguoiChinhSua": session.get('username', 'system')
                    }

                if row_id is not None and row_id in db_ids:
                    # --- UPDATE D√íNG HI·ªÜN C√ì ---
                    update_stmt = sqlalchemy.text("""
                        UPDATE LenhSanXuat_ChiTiet SET
                            STT = :STT, ThoiGianSX = :ThoiGianSX, KichCo = :KichCo, MacThep = :MacThep, SanLuong_1A = :SanLuong_1A,
                            SanLuong_1B = :SanLuong_1B, SanLuong_YeuCau_Cuon = :SanLuong_YeuCau_Cuon, DungSai = :DungSai,
                            CoTinh_GHC = :CoTinh_GHC, CoTinh_GHB = :CoTinh_GHB, CoTinh_GianDai = :CoTinh_GianDai,
                            CoTinh_DoCung = :CoTinh_DoCung, Phoi_MacPhoi = :Phoi_MacPhoi, Phoi_KichThuoc = :Phoi_KichThuoc,
                            YeuCauDacBiet = :YeuCauDacBiet, OrderNumber = :OrderNumber, Batch = :Batch, KL_Cuon = :KL_Cuon,
                            MucDichSuDung = :MucDichSuDung, KhachHang = :KhachHang, DotSX = :DotSX,
                            DaChinhSua = 1, NguoiChinhSua = :NguoiChinhSua, ThoiGianCapNhat = GETDATE()
                        WHERE ID = :ID
                    """)
                    params['ID'] = row_id
                    conn.execute(update_stmt, params)
                else:
                    # --- INSERT D√íNG M·ªöI ---
                    insert_stmt = sqlalchemy.text("""
                        INSERT INTO LenhSanXuat_ChiTiet (
                            STT, ThoiGianSX, KichCo, MacThep, SanLuong_1A, SanLuong_1B, SanLuong_YeuCau_Cuon, DungSai,
                            CoTinh_GHC, CoTinh_GHB, CoTinh_GianDai, CoTinh_DoCung, Phoi_MacPhoi, Phoi_KichThuoc,
                            YeuCauDacBiet, OrderNumber, Batch, KL_Cuon, MucDichSuDung, KhachHang, DotSX,
                            DaChinhSua, NguoiChinhSua, ThoiGianCapNhat
                        ) VALUES (
                            :STT, :ThoiGianSX, :KichCo, :MacThep, :SanLuong_1A, :SanLuong_1B, :SanLuong_YeuCau_Cuon, :DungSai,
                            :CoTinh_GHC, :CoTinh_GHB, :CoTinh_GianDai, :CoTinh_DoCung, :Phoi_MacPhoi, :Phoi_KichThuoc,
                            :YeuCauDacBiet, :OrderNumber, :Batch, :KL_Cuon, :MucDichSuDung, :KhachHang, :DotSX,
                            1, :NguoiChinhSua, GETDATE()
                        )
                    """)
                    conn.execute(insert_stmt, params)
            
        return jsonify({"status": "success", "message": "C·∫≠p nh·∫≠t th√†nh c√¥ng!"})
        
    except Exception as e:
        # Transaction s·∫Ω t·ª± ƒë·ªông rollback khi tho√°t kh·ªèi kh·ªëi 'with' n·∫øu c√≥ l·ªói
        print(f"L·ªói save_lsx_data: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500
from openpyxl.cell.cell import MergedCell
# =================================================================
# API 3: Xu·∫•t Excel (POST)
# (Route n√†y kh√¥ng c·∫ßn 'engine')
# =================================================================
def parse_thg_sx(input_text, current_year):

    try:
        parts = input_text.split('-')
        if len(parts) != 2: 
            return None, None

        part1_raw, part2_raw = parts[0].strip(), parts[1].strip()

        # 1. Ph√¢n t√≠ch ng√†y k·∫øt th√∫c (lu√¥n c√≥ th√°ng, vd: '6/11')
        end_parts = part2_raw.split('/')
        if len(end_parts) != 2: 
            return None, None
        end_day = int(end_parts[0])
        end_month = int(end_parts[1])

        # 2. Ph√¢n t√≠ch ng√†y b·∫Øt ƒë·∫ßu
        start_parts = part1_raw.split('/')
        if len(start_parts) == 2:
            # D·∫°ng '30/10'
            start_day = int(start_parts[0])
            start_month = int(start_parts[1])
        elif len(start_parts) == 1:
            # D·∫°ng '30' (c√πng th√°ng ho·∫∑c th√°ng tr∆∞·ªõc)
            start_day = int(start_parts[0])
            if start_day > end_day:
                # N·∫øu ng√†y b·∫Øt ƒë·∫ßu > ng√†y k·∫øt th√∫c (vd: 30 > 6)
                # -> hi·ªÉu l√† th√°ng tr∆∞·ªõc
                start_month = end_month - 1
                if start_month == 0: 
                    start_month = 12 # X·ª≠ l√Ω tr∆∞·ªùng h·ª£p th√°ng 1
            else:
                # N·∫øu ng√†y b·∫Øt ƒë·∫ßu <= ng√†y k·∫øt th√∫c (vd: 5-10/11)
                # -> hi·ªÉu l√† c√πng th√°ng
                start_month = end_month
        else:
            return None, None

        # 3. ƒê·ªãnh d·∫°ng chu·ªói (th√™m zero-padding v√† nƒÉm)
        start_date_str = f"{start_day:02d}/{start_month:02d}/{current_year}"
        end_date_str = f"{end_day:02d}/{end_month:02d}/{current_year}"
        
        return start_date_str, end_date_str
        
    except Exception as e:
        print(f"--- L·ªñI parse_thg_sx: {e} ---")
        return None, None
from copy import copy
from openpyxl.styles import Alignment
def get_sheet_name_from_dotsx(dot_sx, year):
    """
    S·ª¨A L·ªñI:
    Chuy·ªÉn 'T11 1b' th√†nh '01.11.2025'
    Chuy·ªÉn 'T10 5b' th√†nh '05.10.2025'
    """
    try:
        if not dot_sx or not isinstance(dot_sx, str):
            return f"Sheet_{year}"

        parts = dot_sx.split()
        
        # 1. X·ª≠ l√Ω Th√°ng (lu√¥n ·ªü ph·∫ßn 1)
        month_part = parts[0] # vd: "T10"
        month_str = month_part.replace('T', '').strip()
        month_padded = month_str.zfill(2) # vd: "10"

        # 2. X·ª≠ l√Ω Ng√†y (·ªü ph·∫ßn 2)
        day_padded = "01" # M·∫∑c ƒë·ªãnh l√† 01
        
        if len(parts) > 1:
            day_part_raw = parts[1] # vd: "5b"
            
            # T√¨m s·ªë ƒë·∫ßu ti√™n trong chu·ªói n√†y
            # re.search(r"^\d+", ...) t√¨m c√°c ch·ªØ s·ªë ·ªü ƒê·∫¶U chu·ªói
            match = re.search(r"^\d+", day_part_raw) 
            
            if match:
                day_str = match.group(0) # vd: "5"
                day_padded = day_str.zfill(2) # vd: "05"

        # 3. T·∫°o t√™n
        new_name = f"{day_padded}.{month_padded}.{year}"
        return new_name
        
    except Exception as e:
        print(f"--- L·ªñI get_sheet_name_from_dotsx: {e}. Input: {dot_sx} ---")
        # T√™n d·ª± ph√≤ng n·∫øu l·ªói
        safe_name = re.sub(r'[\[\]\*\/\\?\:]', '_', dot_sx)
        return safe_name[:31]
def copy_sheet_between_workbooks(source_sheet, target_workbook, new_title):
    """
    H√†m copy th·ªß c√¥ng 1 sheet (bao g·ªìm style) t·ª´ workbook n√†y sang workbook kh√°c.
    """
    # 1. T·∫°o sheet m·ªõi trong workbook ƒë√≠ch
    target_sheet = target_workbook.create_sheet(title=new_title)

    # 2. Copy gi√° tr·ªã v√† style c·ªßa t·ª´ng √¥
    for row in source_sheet.iter_rows():
        for cell in row:
            
            # === B·∫ÆT ƒê·∫¶U S·ª¨A L·ªñI ===
            # B·ªè qua c√°c √¥ ƒë√£ b·ªã g·ªôp (MergedCell), 
            # v√¨ ch√∫ng ch·ªâ l√† placeholder. √î g·ªëc (top-left) s·∫Ω ƒë∆∞·ª£c copy.
            if isinstance(cell, MergedCell):
                continue
            # === K·∫æT TH√öC S·ª¨A L·ªñI ===

            new_cell = target_sheet.cell(row=cell.row, column=cell.col_idx, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    # 3. Copy chi·ªÅu cao c·ªßa d√≤ng
    for row_dim_key in source_sheet.row_dimensions:
        if source_sheet.row_dimensions[row_dim_key].height:
            target_sheet.row_dimensions[row_dim_key].height = source_sheet.row_dimensions[row_dim_key].height

    # 4. Copy ƒë·ªô r·ªông c·ªßa c·ªôt
    for col_dim_key in source_sheet.column_dimensions:
        if source_sheet.column_dimensions[col_dim_key].width:
            target_sheet.column_dimensions[col_dim_key].width = source_sheet.column_dimensions[col_dim_key].width
            
    # 5. Copy c√°c √¥ ƒë√£ g·ªôp (merged cells) - Logic n√†y v·∫´n ƒë√∫ng
    for merge_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merge_range))
        
    return target_sheet
@lsx_bp.route('/api/export-with-template', methods=['POST'])
def export_with_template():
    try:
        # --- B·∫ÆT ƒê·∫¶U LOGIC M·ªöI: T·∫†O 1 SHEET DUY NH·∫§T ---
        data_rows = request.get_json()
        if not data_rows:
            return jsonify({"status": "error", "message": "Kh√¥ng c√≥ d·ªØ li·ªáu"}), 400

        # 1. T·∫£i tr·ª±c ti·∫øp file Template ƒë·ªÉ l√†m vi·ªác tr√™n ƒë√≥
        TEMPLATE_PATH = "templates/excel/test.xlsx" 
        wb_out = openpyxl.load_workbook(TEMPLATE_PATH)
        
        # L·∫•y sheet ƒëang ho·∫°t ƒë·ªông ƒë·ªÉ ƒëi·ªÅn d·ªØ li·ªáu
        ws = wb_out.active
        ws.title = "LSX Tong Hop"
        # 3. L·∫•y nƒÉm hi·ªán t·∫°i
        current_year = datetime.today().year 

        # 4. X·ª≠ l√Ω Header (A6 v√† B3)
        # --- Logic c·∫≠p nh·∫≠t B3 (Ti√™u ƒë·ªÅ ch√≠nh) ---
        try:
            original_b3_value = str(ws['B3'].value)
            new_b3_value = re.sub(r"Th√°ng \d+/\d{4}", "T·ªïng H·ª£p", original_b3_value)
            ws['B3'].value = new_b3_value
        except Exception as e:
            print(f"--- WARNING: Kh√¥ng th·ªÉ c·∫≠p nh·∫≠t B3. L·ªói: {e} ---")

        # --- Logic c·∫≠p nh·∫≠t A6 (D·∫£i ng√†y t·ªïng h·ª£p) ---
        all_start_dates, all_end_dates = [], []
        for row in data_rows:
            input_thg_sx = row.get('ThoiGianSX')
            if input_thg_sx:
                start_date_str, end_date_str = parse_thg_sx(input_thg_sx, current_year)
                if start_date_str: all_start_dates.append(datetime.strptime(start_date_str, "%d/%m/%Y"))
                if end_date_str: all_end_dates.append(datetime.strptime(end_date_str, "%d/%m/%Y"))
        
        if all_start_dates and all_end_dates:
            min_start_date = min(all_start_dates).strftime("%d/%m/%Y")
            max_end_date = max(all_end_dates).strftime("%d/%m/%Y")
            try:
                original_a6_value = str(ws['A6'].value)
                date_regex = r"\d{2}/\d{2}/\d{4}t"
                dates_to_insert = [min_start_date, max_end_date]
                def replacer(match):
                    if not dates_to_insert: return match.group(0)
                    return dates_to_insert.pop(0)
                new_a6_value = re.sub(r"\d{2}/\d{2}/\d{4}", replacer, original_a6_value, count=2)
                ws['A6'].value = new_a6_value
            except Exception as e:
                print(f"--- WARNING: Kh√¥ng th·ªÉ c·∫≠p nh·∫≠t A6. L·ªói: {e} ---")

        # 5. C·∫•u h√¨nh ghi d·ªØ li·ªáu
        start_row_write = 10 
        start_row_insert = 11
        start_col = 1
        num_records = len(data_rows)

        # 6. Ch√®n th√™m d√≤ng n·∫øu c·∫ßn
        num_rows_to_insert = num_records - 1
        if num_rows_to_insert > 0:
            ws.insert_rows(start_row_insert, amount=num_rows_to_insert)

        # 7. L·∫∑p qua d·ªØ li·ªáu v√† ƒëi·ªÅn v√†o sheet
        for i, row in enumerate(data_rows):
            cur_row = start_row_write + i
            
            # Copy style t·ª´ d√≤ng m·∫´u cho c√°c d√≤ng ƒë∆∞·ª£c ch√®n m·ªõi
            if i > 0:
                template_row = ws[start_row_write]
                for col_idx, template_cell in enumerate(template_row, start=1):
                    new_cell = ws.cell(row=cur_row, column=col_idx)
                    if template_cell.has_style:
                        new_cell.font = copy(template_cell.font)
                        new_cell.border = copy(template_cell.border)
                        new_cell.fill = copy(template_cell.fill)
                        new_cell.number_format = copy(template_cell.number_format)
                        new_cell.protection = copy(template_cell.protection)
                        new_cell.alignment = copy(template_cell.alignment)
                ws.row_dimensions[cur_row].height = ws.row_dimensions[start_row_write].height

            # Ghi d·ªØ li·ªáu v√†o c√°c √¥
            ws.cell(row=cur_row, column=start_col + 0, value=i + 1)
            
            # X·ª≠ l√Ω v√† ghi c·ªôt ThoiGianSX
            original_thg_sx_row = row.get('ThoiGianSX')
            new_thg_sx_value = original_thg_sx_row
            if original_thg_sx_row:
                start_date_row, end_date_row = parse_thg_sx(original_thg_sx_row, current_year)
                if start_date_row and end_date_row:
                    new_thg_sx_value = f"08h00 {start_date_row}\n-\n08h00\n{end_date_row}"
                    cell_to_format = ws.cell(row=cur_row, column=start_col + 1)
                    cell_to_format.alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
            ws.cell(row=cur_row, column=start_col + 1, value=new_thg_sx_value)

            # Ghi c√°c c·ªôt c√≤n l·∫°i
            ws.cell(row=cur_row, column=start_col + 2, value=row.get('KichCo'))
            ws.cell(row=cur_row, column=start_col + 3, value=row.get('MacThep'))
            ws.cell(row=cur_row, column=start_col + 4, value=row.get('SanLuong_1A'))
            ws.cell(row=cur_row, column=start_col + 5, value=row.get('SanLuong_1B'))
            ws.cell(row=cur_row, column=start_col + 6, value=row.get('SanLuong_YeuCau_Cuon'))
            ws.cell(row=cur_row, column=start_col + 7, value=row.get('DungSai'))
            ws.cell(row=cur_row, column=start_col + 8, value=row.get('CoTinh_GHC'))
            ws.cell(row=cur_row, column=start_col + 9, value=row.get('CoTinh_GHB'))
            ws.cell(row=cur_row, column=start_col + 10, value=row.get('CoTinh_GianDai'))
            ws.cell(row=cur_row, column=start_col + 11, value=row.get('CoTinh_DoCung'))
            ws.cell(row=cur_row, column=start_col + 12, value=row.get('Phoi_MacPhoi'))
            ws.cell(row=cur_row, column=start_col + 13, value=row.get('Phoi_KichThuoc'))
            ws.cell(row=cur_row, column=start_col + 14, value=row.get('YeuCauDacBiet'))
            ws.cell(row=cur_row, column=start_col + 15, value=row.get('OrderNumber'))
            ws.cell(row=cur_row, column=start_col + 16, value=row.get('Batch'))
            ws.cell(row=cur_row, column=start_col + 17, value=row.get('KL_Cuon'))
            ws.cell(row=cur_row, column=start_col + 18, value=row.get('MucDichSuDung'))
            ws.cell(row=cur_row, column=start_col + 19, value=row.get('KhachHang'))

        # 8. C·∫≠p nh·∫≠t c√¥ng th·ª©c SUM
        if num_records > 0:
            last_data_row_index = start_row_write + num_records - 1
            total_row_index = last_data_row_index + 1
            
            col_1A_idx, col_1B_idx, col_SLYC_idx = start_col + 4, start_col + 5, start_col + 6
            col_1A_letter, col_1B_letter, col_SLYC_letter = get_column_letter(col_1A_idx), get_column_letter(col_1B_idx), get_column_letter(col_SLYC_idx)
            
            formula_1A = f"=SUM({col_1A_letter}{start_row_write}:{col_1A_letter}{last_data_row_index})"
            formula_1B = f"=SUM({col_1B_letter}{start_row_write}:{col_1B_letter}{last_data_row_index})"
            formula_SLYC = f"=SUM({col_SLYC_letter}{start_row_write}:{col_SLYC_letter}{last_data_row_index})"
            
            ws.cell(row=total_row_index, column=col_1A_idx).value = formula_1A
            ws.cell(row=total_row_index, column=col_1B_idx).value = formula_1B
            ws.cell(row=total_row_index, column=col_SLYC_idx).value = formula_SLYC

        # 9. L∆∞u file v√†o buffer v√† tr·∫£ v·ªÅ
        file_buffer = io.BytesIO()
        wb_out.save(file_buffer) # L∆∞u workbook ƒë√£ ƒë∆∞·ª£c ch·ªânh s·ª≠a
        file_buffer.seek(0)

        return send_file(
            file_buffer,
            as_attachment=True,
            download_name=f'LSX_TongHop_{datetime.now().strftime("%Y%m%d")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    except Exception as e:
        print(f"L·ªói export_with_template: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500
@lsx_bp.route('/lap-lsx-form')
 # Th√™m check quy·ªÅn n·∫øu c·∫ßn
def lap_lsx():
    """
    Render trang L·∫≠p Form LSX b·∫±ng Handsontable.
    """
    # B·∫°n c√≥ th·ªÉ query d·ªØ li·ªáu m·∫∑c ƒë·ªãnh t·ª´ SQL Server ·ªü ƒë√¢y n·∫øu c·∫ßn
    # (N·∫øu kh√¥ng, c·ª© ƒë·ªÉ Handsontable t·ª± fetch qua API nh∆∞ ta ƒë√£ b√†n)
    
    return render_template('laplsx.html')

# =================================================================
# API 4: Import ƒê∆°n H√†ng t·ª´ Excel v√† ghi ƒë√® d·ªØ li·ªáu
# =================================================================
@lsx_bp.route('/api/import-don-hang', methods=['POST'])
@permission_required('manage_lsx')
def import_don_hang():
    # --- S·ª¨A L·ªñI: Nh·∫≠n file tr·ª±c ti·∫øp t·ª´ request ---
    if 'donhang_input_file' not in request.files:
        return jsonify({"status": "error", "message": "Kh√¥ng t√¨m th·∫•y file 'donhang_input_file' trong request."}), 400

    file = request.files['donhang_input_file']

    if file.filename == '':
        return jsonify({"status": "error", "message": "Ch∆∞a ch·ªçn file n√†o."}), 400

    # L∆∞u file v√†o m·ªôt v·ªã tr√≠ t·∫°m th·ªùi ƒë·ªÉ x·ª≠ l√Ω
    temp_dir = os.path.join("uploads", "temp_imports")
    os.makedirs(temp_dir, exist_ok=True)
    unique_filename = f"{uuid4()}_{file.filename}"
    temp_file_path = os.path.join(temp_dir, unique_filename)
    file.save(temp_file_path)

    try:
        # X·ª≠ l√Ω file t·∫°m ƒë·ªÉ l·∫•y d·ªØ li·ªáu
        df_processed = process_create_lsx(temp_file_path)
        if df_processed is None:
            raise ValueError("File kh√¥ng h·ª£p l·ªá ho·∫∑c kh√¥ng c√≥ d·ªØ li·ªáu (h√†m process_create_lsx tr·∫£ v·ªÅ None).")

        # === S·ª¨A L·ªñI 1: Lo·∫°i b·ªè c·ªôt 'ID' kh·ªèi DataFrame ===
        # C·ªôt 'ID' l√† IDENTITY (t·ª± ƒë·ªông tƒÉng) trong DB, kh√¥ng ƒë∆∞·ª£c ch√®n gi√° tr·ªã t∆∞·ªùng minh.
        df_processed = df_processed.drop(columns=['ID'], errors='ignore')

        # === S·ª¨A L·ªñI 2: ƒê·ªãnh nghƒ©a ki·ªÉu d·ªØ li·ªáu cho c√°c c·ªôt (ƒë·∫∑c bi·ªát l√† DotSX) ===
        # ƒêi·ªÅu n√†y gi√∫p tr√°nh l·ªói 'Invalid column name' v√† ƒë·∫£m b·∫£o ki·ªÉu d·ªØ li·ªáu ƒë√∫ng.
        dtype_mapping = {
            'ThoiGianSX': sqlalchemy.types.NVARCHAR(100),
            'KichCo': sqlalchemy.types.NVARCHAR(50),
            'MacThep': sqlalchemy.types.NVARCHAR(50),
            'DungSai': sqlalchemy.types.NVARCHAR(20),
            'CoTinh_GHC': sqlalchemy.types.NVARCHAR(50),
            'CoTinh_GHB': sqlalchemy.types.NVARCHAR(50),
            'CoTinh_GianDai': sqlalchemy.types.NVARCHAR(50),
            'CoTinh_DoCung': sqlalchemy.types.NVARCHAR(50),
            'Phoi_MacPhoi': sqlalchemy.types.NVARCHAR(50),
            'Phoi_KichThuoc': sqlalchemy.types.NVARCHAR(50),
            'YeuCauDacBiet': sqlalchemy.types.NVARCHAR(255),
            'KL_Cuon': sqlalchemy.types.NVARCHAR(50),
            'MucDichSuDung': sqlalchemy.types.NVARCHAR(100),
            'KhachHang': sqlalchemy.types.NVARCHAR(100),
            'DotSX': sqlalchemy.types.NVARCHAR(50),
            'HasWarning': sqlalchemy.types.Boolean # Th√™m c·ªôt DotSX v√†o mapping
        }

        # Ghi d·ªØ li·ªáu v√†o DB
        with engine.begin() as conn:
            conn.execute(text("TRUNCATE TABLE LenhSanXuat_ChiTiet"))
            df_processed.to_sql(
                'LenhSanXuat_ChiTiet', 
                con=conn, 
                if_exists='append',
                index=False,
                dtype=dtype_mapping # S·ª≠ d·ª•ng dtype_mapping ƒë√£ ƒë·ªãnh nghƒ©a
            )
        
        return jsonify({"status": "success", "message": f"Import th√†nh c√¥ng {len(df_processed)} d√≤ng d·ªØ li·ªáu."})

    except Exception as e:
        return jsonify({"status": "error", "message": f"L·ªói khi x√°c nh·∫≠n import: {str(e)}"}), 500
    finally:
        # Lu√¥n x√≥a file t·∫°m sau khi x·ª≠ l√Ω xong (d√π th√†nh c√¥ng hay th·∫•t b·∫°i)
        if os.path.exists(temp_file_path):
            os.remove(temp_file_path)

# =================================================================
# API 5: X√≥a to√†n b·ªô d·ªØ li·ªáu trong b·∫£ng
# =================================================================
@lsx_bp.route('/api/clear-lsx-data', methods=['POST'])
@permission_required('manage_lsx')
def clear_lsx_data():
    try:
        with engine.begin() as conn:
            # D√πng TRUNCATE TABLE ƒë·ªÉ reset b·∫£ng nhanh v√† hi·ªáu qu·∫£
            conn.execute(text("TRUNCATE TABLE LenhSanXuat_ChiTiet"))
        
        return jsonify({"status": "success", "message": "ƒê√£ x√≥a to√†n b·ªô d·ªØ li·ªáu th√†nh c√¥ng."})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": f"ƒê√£ x·∫£y ra l·ªói khi x√≥a d·ªØ li·ªáu: {str(e)}"}), 500
