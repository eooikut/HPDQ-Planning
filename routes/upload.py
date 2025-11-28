from flask import Blueprint, render_template, request, redirect, url_for, current_app, flash
from uuid import uuid4
import os, shutil, time
import pandas as pd
from datetime import datetime
from db import engine
from storage_utils import load_metadata, save_metadata
from ProcessData import (
    process_lsx,
    process_lichtau,
    process_so_details
)
from upsert_dataSAP import save_df_to_db, save_lichtau
from auth.decorator import permission_required
import logging
from threading import Lock

upload_bp = Blueprint("upload", __name__)

# 1. Khởi tạo khóa toàn cục
upload_lock = Lock()

def read_df_from_db(table_name: str) -> pd.DataFrame:
    try:
        return pd.read_sql_table(table_name, con=engine)
    except Exception as e:
        try:
            return pd.read_sql_query(f"SELECT * FROM {table_name}", con=engine)
        except Exception as e2:
            current_app.logger.warning(f"Lỗi khi đọc bảng {table_name}: {e2}")
            return pd.DataFrame()

@upload_bp.route("/upload", methods=["GET", "POST"])
@permission_required('upload_files')
def upload_files():
    if request.method == "POST":
        # ==============================================================================
        # 2. QUAN TRỌNG: Thử đóng khóa (Kiểm tra xem có ai đang upload không)
        # ==============================================================================
        # blocking=False: Nếu khóa đang đóng thì KHÔNG chờ, trả về False luôn để báo lỗi
        if not upload_lock.acquire(blocking=False):
            flash("⚠️ Hệ thống đang bận xử lý file của người khác. Vui lòng đợi vài phút rồi thử lại!", "warning")
            return redirect(request.url)

        # Nếu code chạy đến đây, nghĩa là mình đã cầm được khóa (Lock acquired)
        try:
            lsx_file = request.files.get("lsx_file")
            lichtau_file = request.files.get("lichtau_file")
            so_detail_files = request.files.getlist("so_detail_files") 

            # Kiểm tra xem có file nào được upload không
            if not lsx_file and not lichtau_file and not any(f for f in so_detail_files):
                flash("Bạn chưa chọn file nào để upload!", "danger")
                return redirect(request.url)

            temp_id = str(uuid4())
            temp_folder = os.path.join(current_app.config["UPLOAD_FOLDER"], f"temp_{temp_id}")
            os.makedirs(temp_folder, exist_ok=True)

            # --- Tạo file log riêng ---
            log_folder = os.path.join(current_app.config["UPLOAD_FOLDER"], "logs")
            os.makedirs(log_folder, exist_ok=True)
            log_path = os.path.join(log_folder, f"upload_{temp_id}.txt")
            log_lines = []
            
            # Import ở đây để tránh lỗi circular import nếu có
            from phanbodudoan import ExportDataSAP
            
            try:
                # --- Xử lý LSX ---
                if lsx_file:
                    try:
                        lsx_filename = os.path.basename(lsx_file.filename).replace(" ", "_")
                        lsx_path = os.path.join(temp_folder, lsx_filename).replace("\\", "/")
                        lsx_file.save(lsx_path)

                        from openpyxl import load_workbook
                        wb = load_workbook(lsx_path, data_only=True)
                        ws = wb.worksheets[3] if len(wb.worksheets) > 3 else wb.active
                        name = ws["B3"].value or os.path.splitext(lsx_filename)[0]

                        metadata = load_metadata()
                        if any(entry.get("name") == name for entry in metadata):
                            msg = f"❌ LSX '{lsx_filename}': lệnh sản xuất '{name}' đã tồn tại"
                            log_lines.append(msg)
                            flash(msg, "danger")
                        else:
                            df_lsx = process_lsx(lsx_path)
                            df_lsx["lsx_id"] = temp_id
                            save_df_to_db(df_lsx, "lsx", engine)
                            metadata.append({
                                "type": "lsx",
                                "id": temp_id,
                                "name": name,
                                "lsx": os.path.join(temp_folder, lsx_filename).replace("\\", "/"),
                                "uploaded_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            })
                            save_metadata(metadata)
                            ExportDataSAP()
                            msg = f"✅ LSX '{lsx_filename}' upload thành công"
                            log_lines.append(msg)
                            flash(msg, "success")
                    except Exception as e:
                        msg = f"❌ LSX '{lsx_file.filename}' lỗi: {e}"
                        log_lines.append(msg)
                        flash(msg, "danger")
                        current_app.logger.exception(e)

                # --- Xử lý Lịch tàu ---
                if lichtau_file:
                    try:
                        lichtau_filename = os.path.basename(lichtau_file.filename).replace(" ", "_")
                        lichtau_path = os.path.join(temp_folder, lichtau_filename).replace("\\", "/")
                        lichtau_file.save(lichtau_path)

                        df_lichtau = process_lichtau(lichtau_path)
                        save_lichtau(df_lichtau, "lichtau", engine)
                        metadata = load_metadata()
                        metadata.append({
                            "type": "lichtau",
                            "tau_name": lichtau_filename,
                            "uploaded_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "file": os.path.join(temp_folder, lichtau_filename).replace("\\", "/"),
                        })
                        save_metadata(metadata)
                        ExportDataSAP()

                        msg = f"✅ Lịch tàu '{lichtau_filename}' upload thành công!!"
                        log_lines.append(msg)
                        flash(msg, "success")
                    except Exception as e:
                        msg = f"❌ Lịch tàu '{lichtau_file.filename}' lỗi: {e}"
                        log_lines.append(msg)
                        flash(msg, "danger")
                        current_app.logger.exception(e)

                # --- Xử lý Chi tiết SO ---
                so_detail_paths = []
                try:
                    for so_file in so_detail_files:
                        if so_file and so_file.filename:
                            filename = os.path.basename(so_file.filename).replace(" ", "_")
                            file_path = os.path.join(temp_folder, filename).replace("\\", "/")
                            so_file.save(file_path)
                            so_detail_paths.append(file_path)
                            metadata = load_metadata()
                            metadata.append({
                                "type": "donhangchitiet",
                                "tau_name": filename, 
                                "uploaded_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                "file": file_path,
                            })
                            log_lines.append(f"✅ Đã lưu tạm file chi tiết SO: {filename}")
                            save_metadata(metadata)
                            ExportDataSAP()
                    if so_detail_paths:
                        process_so_details(so_detail_paths)
                        msg = "✅ Xử lý và cập nhật dữ liệu chi tiết SO thành công!"
                        log_lines.append(msg)
                        flash(msg, "success")
                except Exception as e:
                    msg = f"❌ Lỗi khi xử lý file chi tiết SO: {e}"
                    log_lines.append(msg)
                    flash(msg, "danger")
                    current_app.logger.exception(e)
                
                # --- Chuyển folder tạm -> chính ---
                folder = os.path.join(current_app.config["UPLOAD_FOLDER"], temp_id)
                if os.path.exists(folder):
                    shutil.rmtree(folder)
                os.rename(temp_folder, folder)

            except Exception as e:
                if os.path.exists(temp_folder):
                    shutil.rmtree(temp_folder)
                msg = f"❌ Lỗi chung khi upload: {e}"
                log_lines.append(msg)
                flash(msg, "danger")
                current_app.logger.exception(e)
            
            # Ghi log
            if 'log_path' in locals():
                with open(log_path, "w", encoding="utf-8") as f:
                    f.write("\n".join(log_lines))

        finally:
            # ==============================================================================
            # 3. QUAN TRỌNG: Trả lại chìa khóa để người khác dùng
            # ==============================================================================
            upload_lock.release()

        return redirect(url_for("tau_bp.lichtau"))

    return render_template("upload.html")